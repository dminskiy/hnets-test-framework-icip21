# GOAL:
# Priors: dataset, datasize
# results = {
#               Experiment1 (100% data):{
#                                           Net1: {
#                                                       results: []
#                                                   }
#                                           NetN: {
#                                                       results: []
#                                                   }
#                                       }
#               Experiment2 (75% data): {
#                                           Net1: {
#                                                   results: []
#                                                   }
#                                           NetN: {
#                                                   results: []
#                                                  }
#                                       }
#               ExperimentN
#           }
#
# Save results as JSON
# Save results as Excel
# Get graphs from results

from results_decoder.utils.results_processing_utils import *
from results_decoder.utils.ReportDataHolder import ReportData
from results_decoder.utils.radar_factory import radar_factory

import matplotlib.pyplot as plt
from openpyxl import Workbook
import numpy as np

import datetime
import json
import argparse
import math

MIN_EPOCHS = 120

EXCEL_SUFFIX = ".xlsx"
JSON_SUFFIX = ".json"
JPEG_SUFFIX = ".jpeg"

FINAL_JSON_STD_ENDING = ".DATA_FINAL.json"
STD_DATA_CHECKPT_DIR = os.path.join("CHECKPTS", "data")

FINAL_JSON_INFO_TMPL = {
    "dir":None,
    "filename":None
}

K_MAP = {'8':["wrnscat_12_8", "wrn_16_8"],
         '16':["wrnscat_12_16", "wrn_16_16"]
         }

NET_MAP = {
    'REF':{
        'isHybrid': False,
        'type': None,
        'order': None
    },
    'D':{
        'isHybrid': True,
        'type': 'dtcwt',
        'order': 2
    },
    'Dli':{
        'isHybrid': True,
        'type': 'dtcwt',
        'order': 1
    },
    'DL':{
        'isHybrid': True,
        'type': 'dtcwt_l',
        'order': 2
    },
    'DliL':{
        'isHybrid': True,
        'type': 'dtcwt_l',
        'order': 1
    },
    'M':{
        'isHybrid': True,
        'type': 'mallat',
        'order': 2
    },
    'Mli':{
        'isHybrid': True,
        'type': 'mallat',
        'order': 1
    },
    'ML':{
        'isHybrid': True,
        'type': 'mallat_l',
        'order': 2
    },
    'MliL':{
        'isHybrid': True,
        'type': 'mallat_l',
        'order': 1
    }
}

def drange(x, y, jump):
    out = [x]
    _x = float(x)
    _y = float(y)
    while _x < _y:
        _x += jump
        out.append(round(_x,2))
    return out

def get_net_label(isHybrid, type, order):

    for label in NET_MAP:
        if NET_MAP[label]["isHybrid"] == isHybrid and NET_MAP[label]["type"] == type and NET_MAP[label]["order"] == order:
            return label
    return None


def is_hybrid(report_data):
    return report_data.scat_enabled


def get_report_data(dir):

    final_json_file = get_file(dir, FINAL_JSON_STD_ENDING)

    final_json_info = FINAL_JSON_INFO_TMPL.copy()
    report_data = None

    final_json_info["filename"] = final_json_file
    final_json_info["dir"] = dir

    if final_json_file is None:
        data_checkpoints_dir = os.path.join(dir, STD_DATA_CHECKPT_DIR)
        data_checkpoint = get_file(data_checkpoints_dir, ".json")
        if data_checkpoint is None:
            return report_data, final_json_info
        else:
            final_json_info["filename"] = data_checkpoint
            final_json_info["dir"] = data_checkpoints_dir

    if final_json_info["filename"] is not None:
        json_path = os.path.join(final_json_info["dir"],final_json_info["filename"])
        all_data = load_json(json_path)
        report_data = ReportData()
        report_data.load_report_data(all_data)

    return report_data, final_json_info


class Experiments:
    def __init__(self):
        # List of class Experiment
        self.list = []
        self.dict = None

    def add_experitment(self, experiment):
        self.list.append(experiment)

    def save_dict_as_json(self, out_path, out_name):
        if os.path.isdir(out_path):
            out_file = os.path.join(out_path, out_name)
        else:
            warnings.warn(NotADirectoryError("Can't create table. THe saving directory doesn't exist: [{}]\n"
                                             .format(out_path)))
            return -1

        if self.dict is not None:
            with open(out_file, 'w') as outfile:
                json.dump(self.dict, outfile)

    def create_dict(self):
        self.dict = {}
        exper_count = 0
        self.dict["experiments_number"] = len(self.list)

        for experiment in self.list:
            pri_key = "Experiment" + str(exper_count)
            self.dict[pri_key] = {}

            self.dict[pri_key]["nets_number"] = len(experiment.networks)

            self.dict[pri_key]["input_size"] = experiment.input_size
            self.dict[pri_key]["data_fraction"] = experiment.data_fraction
            self.dict[pri_key]["device"] = experiment.device
            self.dict[pri_key]["num_devices"] = experiment.num_devices
            self.dict[pri_key]["dataset"] = experiment.dataset

            exper_count += 1
            net_count = 0
            for net in experiment.networks:
                net_key = "Net" + str(net_count)
                self.dict[pri_key][net_key] = {}

                self.dict[pri_key][net_key]["optimiser"] = net.optimiser
                self.dict[pri_key][net_key]["scheduler"] = net.scheduler
                self.dict[pri_key][net_key]["scheduler_step"] = net.scheduler_step
                self.dict[pri_key][net_key]["batch_size"] = net.batch_size
                self.dict[pri_key][net_key]["epochs"] = net.epochs
                self.dict[pri_key][net_key]["cnn"] = net.cnn

                self.dict[pri_key][net_key]["isHybrid"] = net.isHybrid
                if net.isHybrid:
                    self.dict[pri_key][net_key]["order"] = net.order
                    self.dict[pri_key][net_key]["type"] = net.type
                    self.dict[pri_key][net_key]["J"] = net.J

                net_count += 1
                result_count = 0
                top1s = []
                top5s = []
                exec_times = []
                train_times = []
                test_times = []
                for result in net.results:
                    result_key = "Result"+str(result_count)
                    self.dict[pri_key][net_key][result_key] = {}

                    self.dict[pri_key][net_key][result_key]["top1"] = result.top1
                    self.dict[pri_key][net_key][result_key]["top5"] = result.top5
                    self.dict[pri_key][net_key][result_key]["original_file"] = result.original_file

                    exec_times.append(result.execution_time)
                    train_times.append(result.total_tst_time/net.epochs)
                    test_times.append(result.total_tr_time/net.epochs)

                    top1s.append(result.top1)
                    top5s.append(result.top5)
                    result_count += 1

                self.dict[pri_key][net_key]["results_number"] = result_count
                self.dict[pri_key][net_key]["average_top1"] = round(np.mean(np.asarray(top1s)), 2)
                self.dict[pri_key][net_key]["average_top5"] = round(np.mean(np.asarray(top5s)), 2)
                self.dict[pri_key][net_key]["stddev_top1"] = round(np.std(np.asarray(top1s)), 2)
                self.dict[pri_key][net_key]["stddev_top5"] = round(np.std(np.asarray(top5s)), 2)

                self.dict[pri_key][net_key]["average_exec_time"] = round(np.mean(np.asarray(exec_times)))
                self.dict[pri_key][net_key]["average_tr_time"] = round(np.mean(np.asarray(train_times)),2)
                self.dict[pri_key][net_key]["average_tst_time"] = round(np.mean(np.asarray(test_times)),2)

    def dict_to_table(self, out_path, out_name):
        if os.path.isdir(out_path):
            out_file = os.path.join(out_path, out_name)
        else:
            warnings.warn(NotADirectoryError("Can't create table. THe saving directory doesn't exist: [{}]\n"
                                             .format(out_path)))
            return -1

        starting_col = 2

        workbook = Workbook()
        for exp_num in range(self.dict["experiments_number"]):
            current_row = 2
            summary = {"Dataset": None,
                       "Data Used": None,
                       "Epochs Trained": None,
                       "Batch Size": None,
                       "Input size": None,
                       "Network": None,
                       "ScatNet": None,
                       "ScatNet Order": None,
                       "Scat J": None,
                       "Number of Runs": None,
                       "Top-1 Average, %": None,
                       "Top-1 Std Dev": None,
                       "Top-5 average, %": None,
                       "Top-5 Std Dev": None,
                       "Av Execution time, hrs": None,
                       "Av Test time, sec": None,
                       "Av Train time, sec": None,
                       "Number of Devices": None,
                       "Device Name(s)": None,
                       "Optimiser": None,
                       "Scheduler Type": None,
                       "Scheduler Step": None,
                       }

            pri_key = "Experiment" + str(exp_num)

            ss_name = str(self.dict[pri_key]["data_fraction"])+"%"
            workbook.create_sheet(ss_name)
            spreadsheet = workbook[ss_name]

            self.add_headings(summary_dict=summary, spreadsheet=spreadsheet, row=current_row, col=starting_col)
            current_row += 1

            summary["Dataset"] = self.dict[pri_key]["dataset"]

            summary["Input size"] = self.dict[pri_key]["input_size"]
            summary["Data Used"] = self.dict[pri_key]["data_fraction"]
            summary["Device Name(s)"] = self.dict[pri_key]["device"]
            try:
                summary["Number of Devices"] = int(self.dict[pri_key]["num_devices"])
            except:
                summary["Number of Devices"] = self.dict[pri_key]["num_devices"]

            for net_num in range(self.dict[pri_key]["nets_number"]):
                net_key = "Net" + str(net_num)

                summary["Optimiser"] = self.dict[pri_key][net_key]["optimiser"]
                summary["Scheduler Type"] = self.dict[pri_key][net_key]["scheduler"]
                summary["Scheduler Step"] = self.dict[pri_key][net_key]["scheduler_step"]
                summary["Batch Size"] = self.dict[pri_key][net_key]["batch_size"]
                summary["Epochs Trained"] = self.dict[pri_key][net_key]["epochs"]
                summary["Network"] = self.dict[pri_key][net_key]["cnn"]

                if self.dict[pri_key][net_key]["isHybrid"]:
                    summary["ScatNet Order"] = int(self.dict[pri_key][net_key]["order"])
                    summary["ScatNet"] = self.dict[pri_key][net_key]["type"]
                    summary["Scat J"] = int(self.dict[pri_key][net_key]["J"])
                else:
                    summary["ScatNet Order"] = "N/A"
                    summary["ScatNet"] = "N/A"
                    summary["Scat J"] = "N/A"

                summary["Number of Runs"] = int(self.dict[pri_key][net_key]["results_number"])
                summary["Top-1 Average, %"] = self.dict[pri_key][net_key]["average_top1"]
                summary["Top-5 average, %"] = self.dict[pri_key][net_key]["average_top5"]
                summary["Top-1 Std Dev"] = self.dict[pri_key][net_key]["stddev_top1"]
                summary["Top-5 Std Dev"] = self.dict[pri_key][net_key]["stddev_top5"]

                summary["Av Execution time, hrs"] = datetime.timedelta(seconds=self.dict[pri_key][net_key]["average_exec_time"])
                summary["Av Train time, sec"] = self.dict[pri_key][net_key]["average_tr_time"]
                summary["Av Test time, sec"] = self.dict[pri_key][net_key]["average_tst_time"]

                values_to_add = self.get_vals_as_list(summary)
                self.add_row(arr=values_to_add, spreadsheet=spreadsheet, row=current_row, col=starting_col)
                current_row += 1

        workbook.remove(workbook['Sheet'])
        workbook.save(out_file)
        workbook.close()

    def dict_to_graph(self, out_dir, out_name, subplots):
        spoke_labels_list = ['M', 'ML', 'Mli', 'MliL', 'D', 'DL', 'Dli', 'DliL' ]
        k8_results_dict = self.map_results(target_k=8)
        k16_results_dict = self.map_results(target_k=16)

        _dataset = self.dict["Experiment0"]["dataset"]

        ks_dicts = {8:k8_results_dict,
                    16:k16_results_dict}

        datas = {}

        for data_partition in k8_results_dict:

            datas[data_partition] = {}

            for key in ks_dicts:
                datas[data_partition][key] = {}
                results_h_list = self.results_dict2list(ks_dicts[key], data_partition, spoke_labels_list)
                results_ref_list = self.results_dict2list_ref(ks_dicts[key], data_partition, len(spoke_labels_list))
                results_max_h_list = [max(results_h_list) for x in range(len(spoke_labels_list))]

                min_floored = round(math.floor(min(min(results_h_list), results_ref_list[0]) / 0.05) * 0.05 - 0.1, 2)
                max_ciel = round(math.floor(max(max(results_h_list), results_ref_list[0]) / 0.05) * 0.05 + 0.1, 2)

                steps = 11
                delta = (max_ciel-min_floored)/steps
                rids_range = drange(min_floored, max_ciel, delta)

                datas[data_partition][key]['rids_range'] = rids_range
                datas[data_partition][key]['data'] = [spoke_labels_list,
                        ('Data,%: {} K: {}\nRef: {}% Best H: {}%'.format(data_partition, key, results_ref_list[0], max(results_h_list)), [
                            results_h_list,
                            results_ref_list,
                            results_max_h_list
                        ])]

        N = len(spoke_labels_list)
        theta = radar_factory(N, frame='circle')

        if subplots == 4:

            fig, axs = plt.subplots(nrows=2, ncols=4, figsize=(20, 12), dpi=150, subplot_kw=dict(projection='radar'))
            plots = {
                    100: {
                        8: axs[0, 0],
                        16: axs[1, 0]
                    },
                    75: {
                        8: axs[0, 1],
                        16: axs[1, 1]
                    },
                    50: {
                        8: axs[0, 2],
                        16: axs[1, 2]
                    },
                    25: {
                        8: axs[0, 3],
                        16: axs[1, 3]
                    },
            }

        if subplots == 3:

            fig, axs = plt.subplots(nrows=2, ncols=3, figsize=(20, 12), dpi=150, subplot_kw=dict(projection='radar'))
            plots = {
                100: {
                    8: axs[0, 0],
                    16: axs[1, 0]
                },
                50: {
                    8: axs[0, 1],
                    16: axs[1, 1]
                },
                25: {
                    8: axs[0, 2],
                    16: axs[1, 2]
                },
            }

        for data_partition in plots:
            for key in plots[data_partition]:
                spoke_labels = datas[data_partition][key]['data'].pop(0)
                title, case_data = datas[data_partition][key]['data'][0]

                plots[data_partition][key].set_rgrids(datas[data_partition][key]['rids_range'])
                plots[data_partition][key].set_title(title, ha='center')

                count = 0
                for d in case_data:
                    if count != 0:
                        plots[data_partition][key].plot(theta, d)
                    else:
                        plots[data_partition][key].plot(theta, d)
                        plots[data_partition][key].fill(theta, d, alpha=0.25)
                    count += 1
                plots[data_partition][key].set_varlabels(spoke_labels)

        fig_title = _dataset.upper()
        fig.suptitle(fig_title, fontsize=20, ha='center')
        fig.tight_layout()
        #plt.show()

        out_name = os.path.join(out_dir, out_name)
        fig.savefig(out_name)

    def results_dict2list_ref(self, results_dict, data_partition, length):
        if 'REF' in results_dict[data_partition]:
            return [results_dict[data_partition]['REF'] for x in range(length)]
        return [100 for x in range(length)]

    def results_dict2list(self, results_dict, data_partition, spoke_labels_list):
        list = []
        for net in spoke_labels_list:
         if net in results_dict[data_partition]:
            list.append(results_dict[data_partition][net])
         else:
            list.append(0)
        return list

    def map_results(self, target_k):
        results = {}

        for exp_num in range(self.dict["experiments_number"]):

            pri_key = "Experiment" + str(exp_num)

            data_fraction = self.dict[pri_key]["data_fraction"]

            results[data_fraction] = {}

            for net_num in range(self.dict[pri_key]["nets_number"]):
                net_key = "Net" + str(net_num)

                net_epochs = self.dict[pri_key][net_key]["epochs"]
                if net_epochs < MIN_EPOCHS:
                    continue

                net = self.dict[pri_key][net_key]["cnn"]
                isHybrid = False
                scat_type = None
                scat_order = None

                if self.dict[pri_key][net_key]["isHybrid"]:
                    isHybrid = True
                    scat_order = int(self.dict[pri_key][net_key]["order"])
                    scat_type = self.dict[pri_key][net_key]["type"]

                top1 = self.dict[pri_key][net_key]["average_top1"]

                nets_k = None
                for k in K_MAP:
                    if net in K_MAP[k]:
                        nets_k = int(k)
                        break

                if nets_k is None:
                    warnings.warn("Can't find k value for the net: [{}]".format(net))
                    continue

                if nets_k == target_k:
                    net_label = get_net_label(isHybrid=isHybrid, type=scat_type, order=scat_order)
                else:
                    continue

                if net_label is None:
                    warnings.warn("Can't find net lables for params (isHm type, order): [{},{},{}]"
                                  .format(isHybrid, scat_type, scat_order))
                    continue

                results[data_fraction][net_label] = top1

        return results


    def get_vals_as_list(self, dictionary):
        vals = []
        for key in dictionary:
            vals.append(dictionary[key])
        return vals

    def add_headings(self, summary_dict, spreadsheet, row, col):
        headings = [key for key in summary_dict]
        self.add_row(arr=headings, spreadsheet=spreadsheet, row=row, col=col)

    def add_row(self, arr, spreadsheet, row, col):
        for i in range(0, len(arr)):
            spreadsheet.cell(row=row, column=col+i).value = arr[i]
        pass


class Experiment:
    def __init__(self):
        self.isEmpty = True

        self.input_size = None
        self.data_fraction = None
        self.device = None
        self.num_devices = None
        self.dataset = None
        # list of class Network or Hybrid
        self.networks = []  # list of unique nets

    def init(self, report_data):
        self.input_size = report_data.data_size
        self.data_fraction = report_data.data_partition
        self.device = report_data.device
        self.num_devices = report_data.num_devices
        self.dataset = report_data.dataset
        self.isEmpty = False

    def new_network(self, network):
        for existing_net in self.networks:
            if existing_net.is_identical(network):
                existing_net.add_result(network.results)
                return
        self.networks.append(network)


class Network:
    def __init__(self, report_data, json_full_path):
        # list of class Result
        self.results = []
        self.optimiser = report_data.optimiser
        self.scheduler = report_data.scheduler_type
        self.scheduler_step = report_data.scheduler_step
        self.batch_size = report_data.batch_sz
        self.epochs = report_data.epochs

        self.cnn = report_data.classifier
        self.isHybrid = False
        if report_data.scat_enabled:
            self.isHybrid = True

        self.results.append(Result(report_data=report_data, full_file_path=json_full_path))

    def is_identical(self, other):
        return self.optimiser == other.optimiser and \
               self.scheduler == other.scheduler and \
               self.scheduler_step == other.scheduler_step and \
               self.batch_size == other.batch_size and \
               self.epochs == other.epochs and\
               self.cnn == other.cnn and \
               self.isHybrid == other.isHybrid

    def add_result(self, result):
        self.results += result


class Hybrid(Network):
    def __init__(self, report_data, json_full_path):
        super().__init__(report_data, json_full_path)
        self.order = report_data.scat_order
        self.type = report_data.scat_type
        self.J = report_data.J

    def is_identical(self, other):
        if not other.isHybrid:
            return False

        return Network.is_identical(self, other) and \
               self.order == other.order and \
               self.type == other.type and\
               self.J == other.J


class Result:
    def __init__(self, report_data, full_file_path):
        self.top1 = report_data.final_accuracy
        self.top5 = report_data.top_5_accuracy
        self.original_file = full_file_path
        self.execution_time = report_data.total_execution_time_sec
        self.total_tst_time = report_data.total_tst_time
        self.total_tr_time = report_data.total_tr_time


def parse_input_arguments():
    parser = argparse.ArgumentParser()

    parser.add_argument('-base_dir', default=None, nargs="?", type=str, required=True,
                        help="Directory that contains results directories as sub-directories. ")
    parser.add_argument("-out_dir", default=None, nargs="?", type=str, required=True,
                        help="Directory where the results will be saved.")
    parser.add_argument("-out_name", default=None, nargs="?", required=True, help="Output filename. excel and json"
                                                                                   " files will be created with"
                                                                                   " this filename")
    parser.add_argument("-exclude_dirs", default=None, nargs="*", required=False, help="Specify directory names to be "
                                                                                       "exlcuded when looking for "
                                                                                       "results directories. Should be used"
                                                                                       "with -results_top_dir to have an effect")

    return parser.parse_args()

if __name__ == '__main__':

    args = parse_input_arguments()

    base_dir = args.base_dir
    out_dir = args.out_dir
    out_name = args.out_name
    exclude_list = args.exclude_dirs

    ensure_dir_exists(out_dir)

    experiments = Experiments()

    empty_dirs = []

    # Count valid subdirs
    sub_dirs = 0
    for subdir in os.listdir(base_dir):
        if not subdir in exclude_list:
            sub_dirs += 1

    # Get list of experiments per subdir
    # Subdir: 25%, 50%, 75%, 100%
    for subdir in os.listdir(base_dir):
        if subdir in exclude_list:
            continue

        current_experiment = Experiment()
        full_subdir = os.path.join(base_dir, subdir)
        for results_folder in os.listdir(full_subdir):

            if results_folder in exclude_list:
                continue

            json_dir = os.path.join(base_dir, subdir, results_folder)

            report_data, json_info = get_report_data(json_dir)

            if report_data is None:
                empty_dirs.append(json_dir)
                warnings.warn("Cannot find JSON file for dir {}".format(json_dir))
                continue

            if current_experiment.isEmpty:
                current_experiment.init(report_data=report_data)

            if is_hybrid(report_data=report_data):
                net = Hybrid(report_data=report_data, json_full_path=os.path.join(json_info["dir"],
                                                                                  json_info["filename"]))
            else:
                net = Network(report_data=report_data, json_full_path=os.path.join(json_info["dir"],
                                                                                   json_info["filename"]))

            current_experiment.new_network(network=net)

        experiments.add_experitment(current_experiment)

    experiments.create_dict()
    experiments.dict_to_table(out_path=out_dir, out_name=out_name+EXCEL_SUFFIX)
    experiments.dict_to_graph(out_dir=out_dir, out_name=out_name+JPEG_SUFFIX, subplots=sub_dirs)
    experiments.save_dict_as_json(out_path=out_dir, out_name=out_name+JSON_SUFFIX)

